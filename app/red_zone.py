from pathlib import Path
from datetime import datetime

import openpyxl

from app.data import df_period1, df_period2
from app.assumption_setting import (
    _compute_cohort,
    OUR_TO_PROPHET,
    CHANNELS,
    PAY_MAP,
    CURRENT_ASSUMPTION,
    OUTPUT_DIR
)

DEVIATION_THRESHOLD = 0.02
DURATION_TO_CHECK = 1


def _cohort_persistency(frame, era, channel, pay_type, duration=None):
    sub = frame[
        (frame["ERA"] == era) &
        (frame["Channel"] == channel) &
        (frame["Short_Long_Pay"] == pay_type)
    ]

    if duration is not None:
        sub = sub[sub["Duration"] == duration]

    numerator = sub["Numerator"].sum()
    denominator = sub["Denominator"].sum()

    if denominator == 0:
        return None

    return numerator / denominator


def identify_red_zone():

    flagged_results = []
    flagged_keys = set()

    for era in OUR_TO_PROPHET:
        for channel in CHANNELS:
            for pay_type in PAY_MAP:

                by_duration, _ = _compute_cohort(era, channel, pay_type)
                proposed_assumption = by_duration.get(DURATION_TO_CHECK)

                latest_actual = _cohort_persistency(df_period2, era, channel, pay_type, duration=DURATION_TO_CHECK)

                if proposed_assumption is None or latest_actual is None:
                    continue

                deviation = round(latest_actual - proposed_assumption, 4)

                if abs(deviation) <= DEVIATION_THRESHOLD:
                    continue

                zone = "red" if deviation < 0 else "green"

                flagged_keys.add((era, channel, pay_type))

                flagged_results.append({
                    "era": era,
                    "channel": channel,
                    "pay_type": pay_type,
                    "duration": DURATION_TO_CHECK,
                    "latest_actual": round(latest_actual, 4),
                    "proposed_assumption": round(proposed_assumption, 4),
                    "deviation": deviation,
                    "zone": zone
                })

    prophet_path = _write_red_zone_prophet_file(flagged_keys)

    return {
        "status": "success",
        "threshold": DEVIATION_THRESHOLD,
        "duration_checked": DURATION_TO_CHECK,
        "flagged_count": len(flagged_results),
        "prophet_file": prophet_path.name,
        "results": flagged_results
    }


def _write_red_zone_prophet_file(flagged_keys, label="RedZone_Update"):
    """
    Builds the full 56-cohort Prophet lapse table. Cohorts in
    flagged_keys get their newly computed proposed assumption; every
    other cohort is held at the current/prior assumption, unchanged.
    """

    OUTPUT_DIR.mkdir(exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BE_Lapse_table"

    ws["F1"] = "Duration >>>>"
    ws["B3"] = "DIST_CHANNEL"
    ws["C3"] = "PPT"
    ws["D3"] = "VARIABLE"
    ws["E3"] = "BE_LAPSE_GRP"
    ws["F3"] = 1

    for col_idx in range(7, 106):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        prev_letter = openpyxl.utils.get_column_letter(col_idx - 1)
        ws[f"{col_letter}3"] = f"={prev_letter}3+1"

    row_num = 4

    for era, era_prophet in OUR_TO_PROPHET.items():
        for channel in CHANNELS:
            for pay_type, pay_prophet in PAY_MAP.items():

                is_flagged = (era, channel, pay_type) in flagged_keys

                if is_flagged:
                    by_duration, ultimate = _compute_cohort(era, channel, pay_type)
                    duration_1_9 = [by_duration.get(d) for d in range(1, 10)]
                else:
                    prior_curve = CURRENT_ASSUMPTION[pay_type]
                    duration_1_9 = prior_curve[:9]
                    ultimate = prior_curve[9]

                if ultimate is None or any(v is None for v in duration_1_9):
                    continue

                ws.cell(row=row_num, column=2, value=channel)
                ws.cell(row=row_num, column=3, value=pay_prophet)
                ws.cell(row=row_num, column=4, value="ANN_LAPSE_PC")
                ws.cell(row=row_num, column=5, value=era_prophet)

                for i, persistency_val in enumerate(duration_1_9):
                    lapse_val = round(1 - persistency_val, 4)
                    col = 6 + i
                    cell = ws.cell(row=row_num, column=col, value=lapse_val)
                    cell.number_format = "0.00%"

                ultimate_lapse = round(1 - ultimate, 4)
                for col in range(15, 106):
                    cell = ws.cell(row=row_num, column=col, value=ultimate_lapse)
                    cell.number_format = "0.00%"

                ws.cell(row=row_num, column=106, value="CHANGED" if is_flagged else "UNCHANGED")

                row_num += 1

    ws.column_dimensions["E"].width = 34
    for col_idx in range(6, 107):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 9

    filename = f"Prophet_{label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = OUTPUT_DIR / filename
    wb.save(file_path)

    return file_path


def generate_single_cohort_prophet_file(era, channel, pay_type):
    """
    Builds the full 56-cohort Prophet table where ONLY this one cohort
    is updated to its proposed assumption — every other cohort, including
    other flagged ones, stays at the current/prior assumption. Used for
    isolating the financial impact of a single cohort change in Prophet.
    """

    target_key = (era, channel, pay_type)
    pay_type_safe = pay_type.replace(" ", "_")
    label = f"{era}_{channel}_{pay_type_safe}"

    return _write_red_zone_prophet_file({target_key}, label=label)


def red_zone_product_breakdown(era, channel, pay_type):

    latest_sub = df_period2[
        (df_period2["ERA"] == era) &
        (df_period2["Channel"] == channel) &
        (df_period2["Short_Long_Pay"] == pay_type) &
        (df_period2["Duration"] == DURATION_TO_CHECK)
    ]

    if latest_sub.empty:
        return {"status": "error", "message": "No data found for this cohort."}

    products = sorted(latest_sub["Product_Name"].unique().tolist())
    breakdown = []

    for product in products:
        latest_prod = latest_sub[latest_sub["Product_Name"] == product]
        latest_num = latest_prod["Numerator"].sum()
        latest_den = latest_prod["Denominator"].sum()
        latest_persistency = round((latest_num / latest_den) * 100, 2) if latest_den else None

        prev_sub = df_period1[
            (df_period1["ERA"] == era) &
            (df_period1["Channel"] == channel) &
            (df_period1["Short_Long_Pay"] == pay_type) &
            (df_period1["Duration"] == DURATION_TO_CHECK) &
            (df_period1["Product_Name"] == product)
        ]
        prev_num = prev_sub["Numerator"].sum()
        prev_den = prev_sub["Denominator"].sum()
        previous_persistency = round((prev_num / prev_den) * 100, 2) if prev_den else None

        change = (
            round(latest_persistency - previous_persistency, 2)
            if latest_persistency is not None and previous_persistency is not None
            else None
        )

        breakdown.append({
            "product": product,
            "latest_persistency": latest_persistency,
            "previous_persistency": previous_persistency,
            "change": change
        })

    return {
        "status": "success",
        "era": era,
        "channel": channel,
        "pay_type": pay_type,
        "products": breakdown
    }