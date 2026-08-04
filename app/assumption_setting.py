from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill

from app.data import df_period1, df_period2

# ---- Mapping: our dataset's ERA naming -> Prophet's BE_LAPSE_GRP naming ----
OUR_TO_PROPHET = {
    "Protection_ERA2": "Protection_ERA2",
    "Protection_ERA3": "Protection_ERA3",
    "Savings_Income_Post_PPT_ERA2": "Savings_ERA2_Income_Post_PPT",
    "Savings_Income_Post_PPT_ERA3": "Savings_ERA3_Income_Post_PPT",
    "Savings_Income_within_PPT_ERA3": "Savings_ERA3_Income_Within_PPT",
    "Savings_Lumpsum_ERA2": "Savings_ERA2_Lumpsum",
    "Savings_Lumpsum_ERA3": "Savings_ERA3_Lumpsum",
}

PAY_MAP = {"Short Pay": "SHORT PAY", "Long Pay": "LONG PAY"}
CHANNELS = ["Axis", "Own", "Online", "Others"]

# Demo prior assumption curve (duration 1-9, then a 10+ ultimate value).
# Replace with real per-cohort priors once available.
CURRENT_ASSUMPTION = {
    "Short Pay": [0.50, 0.60, 0.68, 0.74, 0.78, 0.81, 0.83, 0.85, 0.87, 0.88],
    "Long Pay":  [0.55, 0.65, 0.72, 0.78, 0.82, 0.85, 0.87, 0.89, 0.90, 0.92],
}

# Stage 1: blend actual experience across the two periods
EXPERIENCE_WEIGHT_OLD = 0.3
EXPERIENCE_WEIGHT_LATEST = 0.7

# Stage 2: blend that against the prior assumption
FINAL_WEIGHT_PRIOR = 0.5
FINAL_WEIGHT_ACTUAL = 0.5

OUTPUT_DIR = Path(__file__).resolve().parent.parent / "outputs"


def _persistency(frame, era, channel, pay, duration=None, min_duration=None):
    sub = frame[
        (frame["ERA"] == era) &
        (frame["Channel"] == channel) &
        (frame["Short_Long_Pay"] == pay)
    ]

    if duration is not None:
        sub = sub[sub["Duration"] == duration]
    elif min_duration is not None:
        sub = sub[sub["Duration"] >= min_duration]

    numerator = sub["Numerator"].sum()
    denominator = sub["Denominator"].sum()

    if denominator == 0:
        return None

    return numerator / denominator


def _compute_cohort(era_our, channel, pay_our):
    """
    Returns (persistency_by_duration_1_to_9, ultimate_persistency_10plus)
    for a single ERA x Channel x Pay Type cohort.
    """

    prior_curve = CURRENT_ASSUMPTION[pay_our]
    by_duration = {}

    for d in range(1, 10):
        p_old = _persistency(df_period1, era_our, channel, pay_our, duration=d)
        p_new = _persistency(df_period2, era_our, channel, pay_our, duration=d)

        if p_old is None or p_new is None:
            by_duration[d] = None
            continue

        weighted = EXPERIENCE_WEIGHT_OLD * p_old + EXPERIENCE_WEIGHT_LATEST * p_new
        prior = prior_curve[d - 1]
        proposed = FINAL_WEIGHT_PRIOR * prior + FINAL_WEIGHT_ACTUAL * weighted
        by_duration[d] = round(proposed, 4)

    p_old_ult = _persistency(df_period1, era_our, channel, pay_our, min_duration=10)
    p_new_ult = _persistency(df_period2, era_our, channel, pay_our, min_duration=10)

    if p_old_ult is None or p_new_ult is None:
        ultimate = None
    else:
        weighted_ult = EXPERIENCE_WEIGHT_OLD * p_old_ult + EXPERIENCE_WEIGHT_LATEST * p_new_ult
        prior_ult = prior_curve[9]
        ultimate = round(FINAL_WEIGHT_PRIOR * prior_ult + FINAL_WEIGHT_ACTUAL * weighted_ult, 4)

    return by_duration, ultimate


def _compute_cohort_full(era_our, channel, pay_our):
    """
    Same calculation as _compute_cohort, but returns every intermediate
    row (Current Assumption, Oct'25 experience, Jun'26 experience,
    Weighted experience, Proposed assumption) for all 10 duration
    columns (1-9 plus a 10+ ultimate) — used for the Word report, which
    shows the full working, not just the final figure.
    """

    prior_curve = CURRENT_ASSUMPTION[pay_our]
    prior_row, old_row, new_row, weighted_row, proposed_row = [], [], [], [], []

    for d in range(1, 10):
        p_old = _persistency(df_period1, era_our, channel, pay_our, duration=d)
        p_new = _persistency(df_period2, era_our, channel, pay_our, duration=d)

        if p_old is None or p_new is None:
            prior_row.append(None); old_row.append(None); new_row.append(None)
            weighted_row.append(None); proposed_row.append(None)
            continue

        weighted = EXPERIENCE_WEIGHT_OLD * p_old + EXPERIENCE_WEIGHT_LATEST * p_new
        prior = prior_curve[d - 1]
        proposed = FINAL_WEIGHT_PRIOR * prior + FINAL_WEIGHT_ACTUAL * weighted

        prior_row.append(prior)
        old_row.append(p_old)
        new_row.append(p_new)
        weighted_row.append(weighted)
        proposed_row.append(proposed)

    p_old_ult = _persistency(df_period1, era_our, channel, pay_our, min_duration=10)
    p_new_ult = _persistency(df_period2, era_our, channel, pay_our, min_duration=10)

    if p_old_ult is None or p_new_ult is None:
        prior_row.append(None); old_row.append(None); new_row.append(None)
        weighted_row.append(None); proposed_row.append(None)
    else:
        weighted_ult = EXPERIENCE_WEIGHT_OLD * p_old_ult + EXPERIENCE_WEIGHT_LATEST * p_new_ult
        prior_ult = prior_curve[9]
        proposed_ult = FINAL_WEIGHT_PRIOR * prior_ult + FINAL_WEIGHT_ACTUAL * weighted_ult

        prior_row.append(prior_ult)
        old_row.append(p_old_ult)
        new_row.append(p_new_ult)
        weighted_row.append(weighted_ult)
        proposed_row.append(proposed_ult)

    return {
        "Current Assumption": prior_row,
        "YTD Oct'25 Experience": old_row,
        "YTD Jun'26 Experience": new_row,
        "Weighted Experience": weighted_row,
        "Proposed Assumption": proposed_row
    }


def get_full_breakdown():
    """
    Full duration-banded working for every ERA x Channel combination
    (both Pay Types nested inside each block), matching the structure
    of the original assumption-setting calculation template.
    """

    blocks = []

    for era_our in OUR_TO_PROPHET:
        for channel in CHANNELS:
            pay_data = {}
            for pay_our in PAY_MAP:
                pay_data[pay_our] = _compute_cohort_full(era_our, channel, pay_our)

            blocks.append({
                "era": era_our,
                "channel": channel,
                "pay_data": pay_data
            })

    return blocks


def run_assumption_setting():
    """
    Runs the full assumption-setting cycle across every ERA x Channel x
    Short/Long Pay cohort present in the data: blends actual experience
    across both periods (30% older + 70% latest), blends that against the
    prior assumption (50/50), and produces duration-banded proposed
    assumptions (duration 1-9 individually, plus a 10+ ultimate). Writes
    three files: a readable summary workbook, the Prophet-format lapse
    table, and a Word report with a Gemini-written executive summary.
    """

    from app.report_generator import generate_word_report

    cohort_results = []

    for era_our, era_prophet in OUR_TO_PROPHET.items():
        for channel in CHANNELS:
            for pay_our, pay_prophet in PAY_MAP.items():

                by_duration, ultimate = _compute_cohort(era_our, channel, pay_our)

                if ultimate is None:
                    continue

                headline = by_duration.get(1)
                prior_headline = CURRENT_ASSUMPTION[pay_our][0]
                movement = round(headline - prior_headline, 4) if headline is not None else None

                zone = "green" if movement and movement > 0 else "red" if movement and movement < 0 else "neutral"

                cohort_results.append({
                    "era": era_our,
                    "era_prophet": era_prophet,
                    "channel": channel,
                    "pay_type": pay_our,
                    "prior_assumption_duration1": prior_headline,
                    "proposed_assumption_duration1": headline,
                    "proposed_assumption_ultimate": ultimate,
                    "movement": movement,
                    "zone": zone,
                    "by_duration": by_duration
                })

    improved = sum(1 for r in cohort_results if r["zone"] == "green")
    declined = sum(1 for r in cohort_results if r["zone"] == "red")
    unchanged = sum(1 for r in cohort_results if r["zone"] == "neutral")

    prophet_path = _write_prophet_file(cohort_results)
    summary_path = _write_summary_file(cohort_results)

    full_breakdown = get_full_breakdown()
    word_path = generate_word_report(cohort_results, full_breakdown)

    return {
        "status": "success",
        "cohorts_updated": len(cohort_results),
        "improved": improved,
        "declined": declined,
        "unchanged": unchanged,
        "prophet_file": prophet_path.name,
        "summary_file": summary_path.name,
        "word_file": word_path.name,
        "prophet_updates": cohort_results
    }


def _write_prophet_file(cohort_results):
    OUTPUT_DIR.mkdir(exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BE_Lapse_table"

    ws["F1"] = "Duration >>>>"
    ws["B3"] = "DIST_CHANNEL"
    ws["C3"] = "PPT"
    ws["D3"] = "VARIABLE"
    ws["E3"] = "BE_LAPSE_GRP"
    ws["F3"] = 1

    for col_idx in range(7, 106):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        prev_letter = openpyxl.utils.get_column_letter(col_idx - 1)
        ws[f"{col_letter}3"] = f"={prev_letter}3+1"

    row_num = 4
    for r in cohort_results:
        ws.cell(row=row_num, column=2, value=r["channel"])
        ws.cell(row=row_num, column=3, value=PAY_MAP[r["pay_type"]])
        ws.cell(row=row_num, column=4, value="ANN_LAPSE_PC")
        ws.cell(row=row_num, column=5, value=r["era_prophet"])

        for d in range(1, 10):
            persistency_val = r["by_duration"].get(d)
            lapse_val = round(1 - persistency_val, 4) if persistency_val is not None else None
            col = 5 + d
            cell = ws.cell(row=row_num, column=col, value=lapse_val)
            cell.number_format = "0.00%"

        ultimate_lapse = round(1 - r["proposed_assumption_ultimate"], 4)
        for col in range(15, 106):
            cell = ws.cell(row=row_num, column=col, value=ultimate_lapse)
            cell.number_format = "0.00%"

        row_num += 1

    ws.column_dimensions["E"].width = 34
    for col_idx in range(6, 106):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 9

    filename = f"Prophet_BE_Lapse_table_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = OUTPUT_DIR / filename
    wb.save(file_path)

    return file_path


def _write_summary_file(cohort_results):
    OUTPUT_DIR.mkdir(exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assumption Setting Summary"

    headers = [
        "ERA", "Channel", "Pay Type", "Prior Assumption (Dur 1)"
    ] + [f"Proposed Dur {d}" for d in range(1, 10)] + [
        "Proposed Ultimate (10+)", "Movement (Dur 1)", "Zone"
    ]

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10243E", end_color="10243E", fill_type="solid")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    green_fill = PatternFill(start_color="E3F3EA", end_color="E3F3EA", fill_type="solid")
    red_fill = PatternFill(start_color="FAEAE7", end_color="FAEAE7", fill_type="solid")
    neutral_fill = PatternFill(start_color="F0F0EE", end_color="F0F0EE", fill_type="solid")

    row_num = 2
    for r in cohort_results:
        fill = green_fill if r["zone"] == "green" else red_fill if r["zone"] == "red" else neutral_fill

        values = [r["era"], r["channel"], r["pay_type"], r["prior_assumption_duration1"]]
        values += [r["by_duration"].get(d) for d in range(1, 10)]
        values += [r["proposed_assumption_ultimate"], r["movement"], r["zone"].upper()]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = Font(name="Arial")
            cell.fill = fill
            if isinstance(value, float):
                cell.number_format = "0.00%"

        row_num += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    for col_idx in range(4, 17):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 14

    ws.freeze_panes = "A2"

    filename = f"Assumption_Setting_Summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = OUTPUT_DIR / filename
    wb.save(file_path)

    return file_path